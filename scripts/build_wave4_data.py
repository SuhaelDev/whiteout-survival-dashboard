#!/usr/bin/env python3
"""Wave 4 data build: patches data/game-data.json with
- corrected per-pet max levels (pet_levels trimmed per pet)
- pet metadata: rarity, max level, unlock chain, active skill ladder per advancement,
  refinement mark costs per quality tier, passive Troop ATK/DEF anchors per 0.1 step
- pet material chests (custom chest + custom pet chest pack tiers)
- troop training tables (per type / tier: svs points, base seconds, rss)
- expert learning XP rate (60 XP per minute)
- SvS activity point rates (structured)
- new backpack resource types: chief stamina, widgets per hero generation,
  pet custom chests, healing speedups

Sources:
- "Whiteout Survival Calculator v3.1c-rev2 (1).xlsx" (uploads) - troop data, pet costs, svs
- wostools.net pets-calculator data tables (refinement costs, skill ladders, troop AD anchors)
- whiteoutdata.com (pet max levels cross-check, widget level costs)
"""
import json
import sys
import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
GAME_DATA = ROOT / "data" / "game-data.json"
WORKBOOK = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT.parent / "uploads" / "Whiteout Survival Calculator v3.1c-rev2 (1).xlsx"

# ---------------------------------------------------------------- wostools data
# Passive Troop Attack/Defense % anchors. post_adv[k] = value right after k-th
# advancement (level k*10 + .1); pre_adv[k] = value at level (k+1)*10 before advancing.
TROOP_AD = {
    "Common": {"post": [0.05, 0.92, 1.85, 2.87, 3.9, 5.02], "pre": [0.5, 1.42, 2.35, 3.37, 4.4]},
    "N": {"post": [0.08, 1.51, 3.02, 4.7, 6.37, 8.21, 10.06], "pre": [0.84, 2.35, 3.86, 5.53, 7.21, 9.05]},
    "R": {"post": [0.11, 1.92, 3.84, 5.95, 8.08, 10.41, 12.74, 15.08], "pre": [1.08, 2.99, 4.91, 7.03, 9.16, 11.49, 13.82]},
    "SR": {"post": [0.14, 2.37, 4.75, 7.38, 10.01, 12.89, 15.77, 18.65, 21.78], "pre": [1.36, 3.73, 6.12, 8.74, 11.37, 14.25, 17.13, 20.01]},
    "SSR": {"post": [0.17, 2.86, 5.7, 8.85, 11.99, 15.44, 18.88, 22.32, 26.05, 29.78, 33.52], "pre": [1.68, 4.53, 7.38, 10.52, 13.67, 17.12, 20.55, 24.0, 27.72, 31.46]},
}

REFINEMENT = {
    "Common": [("Green", 1, 0), ("Blue", 5, 0), ("Purple", 20, 0), ("Gold", 0, 4)],
    "N": [("Green", 2, 0), ("Blue", 8, 0), ("Purple", 30, 0), ("Gold", 0, 5)],
    "R": [("Green", 3, 0), ("Blue", 12, 0), ("Purple", 40, 0), ("Gold", 0, 7)],
    "SR": [("Green", 4, 0), ("Blue", 15, 0), ("Purple", 50, 0), ("Gold", 0, 8)],
    "SSR": [("Green", 5, 0), ("Blue", 20, 0), ("Purple", 60, 0), ("Gold", 0, 10)],
}

PETS_META = [
    ("cave_hyena", "Common", 50, None, None, "Builder's Aide",
     ["+5% Const. Speed", "+7% Const. Speed", "+9% Const. Speed", "+12% Const. Speed", "+15% Const. Speed"], "23h", "5 min"),
    ("arctic_wolf", "N", 60, "cave_hyena", 15, "Arctic Embrace",
     ["+35 Stamina", "+40 Stamina", "+45 Stamina", "+50 Stamina", "+55 Stamina", "+60 Stamina"], "23h", None),
    ("musk_ox", "N", 60, "arctic_wolf", 15, "Burden Bearer",
     ["CD: 35h", "CD: 31h", "CD: 27h", "CD: 23h", "CD: 19h", "CD: 15h"], "varies", None),
    ("giant_tapir", "R", 70, "musk_ox", 15, "Natural Intuition",
     ["+200 Pet Food", "+250 Pet Food", "+300 Pet Food", "+350 Pet Food", "+400 Pet Food", "+450 Pet Food", "+500 Pet Food"], "23h", None),
    ("titan_roc", "R", 70, "giant_tapir", 15, "Razorbeak",
     ["-1.5% Enemy HP", "-2% Enemy HP", "-2.5% Enemy HP", "-3% Enemy HP", "-3.5% Enemy HP", "-4% Enemy HP", "-5% Enemy HP"], "20h", "2h"),
    ("giant_elk", "SR", 80, "titan_roc", 15, "Mystical Finding",
     ["CD: 51h", "CD: 47h", "CD: 43h", "CD: 39h", "CD: 35h", "CD: 31h", "CD: 27h", "CD: 23h"], "varies", None),
    ("snow_leopard", "SR", 80, "giant_elk", 15, "Lightning Raid",
     ["+15% March, -1.5% Lethal", "+17% March, -2% Lethal", "+19% March, -2.5% Lethal", "+21% March, -3% Lethal",
      "+23% March, -3.5% Lethal", "+25% March, -4% Lethal", "+27% March, -4.5% Lethal", "+30% March, -5% Lethal"], "20h", "2h"),
    ("cave_lion", "SSR", 100, "snow_leopard", 15, "Feral Anthem",
     ["+2.5% Troop ATK", "+3% Troop ATK", "+3.5% Troop ATK", "+4% Troop ATK", "+5% Troop ATK",
      "+6% Troop ATK", "+7% Troop ATK", "+8% Troop ATK", "+9% Troop ATK", "+10% Troop ATK"], "20h", "2h"),
    ("snow_ape", "SSR", 100, "cave_lion", 30, "Tumbling Power",
     ["+1,500 Squad Cap", "+3,000 Squad Cap", "+4,500 Squad Cap", "+6,000 Squad Cap", "+7,500 Squad Cap",
      "+9,000 Squad Cap", "+10,500 Squad Cap", "+12,000 Squad Cap", "+13,500 Squad Cap", "+15,000 Squad Cap"], "20h", "2h"),
    ("iron_rhino", "SSR", 100, "snow_ape", 30, "Rallying Beasts",
     ["+60k Rally Cap", "+70k Rally Cap", "+80k Rally Cap", "+90k Rally Cap", "+100k Rally Cap",
      "+110k Rally Cap", "+120k Rally Cap", "+130k Rally Cap", "+140k Rally Cap", "+150k Rally Cap"], "20h", "2h"),
    ("sabertooth_tiger", "SSR", 100, "iron_rhino", 30, "Apex Assault",
     ["+2.5% Lethality", "+3% Lethality", "+3.5% Lethality", "+4% Lethality", "+5% Lethality",
      "+6% Lethality", "+7% Lethality", "+8% Lethality", "+9% Lethality", "+10% Lethality"], "20h", "2h"),
    ("mammoth", "SSR", 100, "sabertooth_tiger", 30, "Hardened Skin",
     ["+2.5% Troop DEF", "+3% Troop DEF", "+3.5% Troop DEF", "+4% Troop DEF", "+5% Troop DEF",
      "+6% Troop DEF", "+7% Troop DEF", "+8% Troop DEF", "+9% Troop DEF", "+10% Troop DEF"], "20h", "2h"),
    ("frost_gorilla", "SSR", 100, "mammoth", 30, "Earthbound Vigor",
     ["+2.5% Troop HP", "+3% Troop HP", "+3.5% Troop HP", "+4% Troop HP", "+5% Troop HP",
      "+6% Troop HP", "+7% Troop HP", "+8% Troop HP", "+9% Troop HP", "+10% Troop HP"], "20h", "2h"),
    ("frostscale_chameleon", "SSR", 100, "frost_gorilla", 30, "Icy Shroud",
     ["-2.5% Enemy DEF", "-3% Enemy DEF", "-3.5% Enemy DEF", "-4% Enemy DEF", "-5% Enemy DEF",
      "-6% Enemy DEF", "-7% Enemy DEF", "-8% Enemy DEF", "-9% Enemy DEF", "-10% Enemy DEF"], "20h", "2h"),
    ("abyssal_shelldragon", "SSR", 100, "frostscale_chameleon", 30, "Abyssal Skill",
     [], "20h", "2h"),
]

CUSTOM_CHEST_OPTIONS = [
    {"resource": "pet_manuals", "label": "Taming Manual", "per_chest": 7},
    {"resource": "pet_potions", "label": "Energizing Potion", "per_chest": 2},
    {"resource": "pet_serum", "label": "Strengthening Serum", "per_chest": 1},
]

CUSTOM_PET_CHEST_PACK = {
    "note": "Custom Pet Chest Pack - choose 3 rewards from each tier",
    "tiers": [5, 10, 20, 50, 100],
    "options": [
        {"label": "Pet Food", "resource": "pet_food", "amounts": [5000, 10000, 20000, 50000, 100000]},
        {"label": "Pet Mats Chest", "resource": "pet_custom_chests", "amounts": [6, 12, 24, 60, 120]},
        {"label": "Common Wild Marks", "resource": "common_wild_marks", "amounts": [16, 32, 64, 160, 320]},
        {"label": "Advanced Wild Marks", "resource": "advanced_wild_marks", "amounts": [2, 4, 8, 20, 40]},
    ],
}

SVS_POINT_RATES = {
    "fire_crystal": 2000,
    "refined_fire_crystal": 30000,
    "fire_crystal_shard_research": 1000,
    "speedup_minute": 30,
    "book_of_knowledge": 60,
    "expert_sigil": 6000,
    "rare_hero_shard": 350,
    "epic_hero_shard": 1220,
    "mythic_hero_shard": 3040,
    "widget": 8000,
    "essence_stone": 4000,
    "mithril": 144000,
    "lucky_wheel_spin": 8000,
    "common_wild_mark": 1150,
    "advanced_wild_mark": 15000,
    "pet_advancement_point": 50,
    "chief_gear_per_power": 36,
    "chief_charm_per_power": 70,
    "gather_meat_per_1000": 2,
    "gather_wood_per_1000": 2,
    "gather_coal_per_200": 2,
    "gather_iron_per_50": 2,
    "valeria_well_prepared_pct_per_level": 2,
}

WIDGET_LEVEL_COSTS = [5, 10, 15, 20, 25, 30, 35, 40, 45, 50]

EXPERT_XP_PER_LEARNING_MINUTE = 60


def load_pet_columns(wb):
    """Pet Data sheet: per pet column block of Food/Manual/Potion/Serum/SVS."""
    ws = wb["Pet Data"]
    grid = [[c.value for c in row] for row in ws.iter_rows()]
    names_row = grid[3]
    pet_cols = [(i, v) for i, v in enumerate(names_row) if isinstance(v, str) and v.strip()]
    levels = []  # (row_idx, level_code)
    for r in range(5, len(grid)):
        code = grid[r][2]
        if code is None:
            continue
        levels.append((r, code))
    return grid, pet_cols, levels


def rebuild_pet_levels(data, wb):
    grid, pet_cols, levels = load_pet_columns(wb)
    name_to_id = {p["name"]: p["pet_id"] for p in data["pets"]}
    out = []
    max_levels = {}
    for col, name in pet_cols:
        pet_id = name_to_id.get(name.strip())
        if not pet_id:
            continue
        meta = next((m for m in PETS_META if m[0] == pet_id), None)
        cap = meta[2] if meta else 100
        max_levels[pet_id] = cap
        out.append({
            "pet_id": pet_id, "level_code": "0", "food": 0, "manuals": 0,
            "potions": 0, "serum": 0, "svs_points": 0,
            "source_sheet": "Pet Data", "source_row": 6,
        })
        for r, code in levels:
            level_num = float(code)
            if level_num <= 0 or level_num > cap + 0.2:
                continue
            is_adv = abs(level_num - round(level_num)) > 1e-9
            row = grid[r]
            def num(offset):
                v = row[col + offset] if col + offset < len(row) else None
                return float(v) if isinstance(v, (int, float)) else 0
            entry = {
                "pet_id": pet_id,
                "level_code": (str(int(level_num)) if not is_adv else f"{level_num:.1f}"),
                "food": num(0) if not is_adv else 0,
                "manuals": num(2) if is_adv else 0,
                "potions": num(4) if is_adv else 0,
                "serum": num(6) if is_adv else 0,
                "svs_points": num(8) if is_adv else 0,
                "source_sheet": "Pet Data",
                "source_row": r + 1,
            }
            out.append(entry)
    data["pet_levels"] = out
    return max_levels


def enrich_pets(data, max_levels):
    by_id = {p["pet_id"]: p for p in data["pets"]}
    for (pet_id, rarity, cap, prev_id, prev_lv, skill, effects, cooldown, duration) in PETS_META:
        pet = by_id.get(pet_id)
        if not pet:
            continue
        pet["rarity"] = rarity
        pet["max_level"] = max_levels.get(pet_id, cap)
        pet["unlocks_after_pet_id"] = prev_id
        pet["unlocks_after_level"] = prev_lv
        pet["skill_name"] = skill
        pet["skill_effects"] = effects
        pet["skill_cooldown"] = cooldown
        pet["skill_duration"] = duration
        ad = TROOP_AD[rarity]
        pet["troop_ad_post_advancement"] = ad["post"]
        pet["troop_ad_pre_advancement"] = ad["pre"]
        pet["refinement_costs"] = [
            {"tier": tier, "common_wild_marks": c, "advanced_wild_marks": a}
            for tier, c, a in REFINEMENT[rarity]
        ]
    data["pet_material_chest"] = {
        "chest_resource": "pet_custom_chests",
        "options": CUSTOM_CHEST_OPTIONS,
        "note": "Each Pet Advancement Materials Custom Chest opens into ONE choice: 7 Taming Manuals, 2 Energizing Potions or 1 Strengthening Serum.",
    }
    data["custom_pet_chest_pack"] = CUSTOM_PET_CHEST_PACK


def build_troop_tiers(data, wb):
    ws = wb["Troop Data"]
    grid = [[c.value for c in row] for row in ws.iter_rows()]
    blocks = {"Infantry": 3, "Lancer": 20, "Marksman": 37}
    rows = []
    for troop, start in blocks.items():
        for r in range(start, start + 12):
            row = grid[r]
            if not isinstance(row[1], (int, float)):
                continue
            rows.append({
                "troop_type": troop.lower(),
                "tier": int(row[2]),
                "svs_points": float(row[1]),
                "base_seconds": float(row[3]),
                "meat": float(row[5] or 0),
                "wood": float(row[6] or 0),
                "coal": float(row[7] or 0),
                "iron": float(row[8] or 0),
                "source_sheet": "Troop Data",
                "source_row": r + 1,
            })
    data["troop_tiers"] = rows
    data["troop_training_notes"] = {
        "time_formula": "base_seconds / (1 + training_speed_pct/100) per troop",
        "promotion_cost": "cost(target tier) - cost(source tier); time likewise",
        "promotion_points": "svs_points(target) - svs_points(source)",
        "training_points": "svs_points(tier) per troop trained",
        "speedup_points_per_minute": 30,
    }


def main():
    data = json.loads(GAME_DATA.read_text(encoding="utf-8"))
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)

    max_levels = rebuild_pet_levels(data, wb)
    enrich_pets(data, max_levels)
    build_troop_tiers(data, wb)

    data["svs_point_rates"] = SVS_POINT_RATES
    data["expert_learning"] = {
        "xp_per_learning_minute": EXPERT_XP_PER_LEARNING_MINUTE,
        "note": "Expert skill XP accrues at 60 XP per minute of active learning; learning speedups reduce the wall-clock wait for the same XP.",
    }
    data["hero_widget_level_costs"] = WIDGET_LEVEL_COSTS

    # -------- resource types
    existing = {r["resource_id"] for r in data["resource_types"]}
    def add_resource(rid, name, category, unit="quantity"):
        if rid not in existing:
            data["resource_types"].append({
                "resource_id": rid, "name": name, "category": category,
                "unit": unit, "source_sheet": "wave4", "source_row": None,
            })
            existing.add(rid)

    add_resource("chief_stamina", "Chief Stamina", "chief_items")
    add_resource("stamina_cans", "Stamina Cans (10)", "chief_items")
    add_resource("pet_custom_chests", "Pet Adv. Custom Chests", "pet_materials")
    add_resource("common_wild_marks", "Common Wild Marks", "pet_materials")
    add_resource("advanced_wild_marks", "Advanced Wild Marks", "pet_materials")
    add_resource("healing_speedups_minutes", "Healing Speedups", "speedups", "minutes")

    gens = sorted({int(h["generation"]) for h in data["heroes"] if h.get("generation")})
    for g in gens:
        add_resource(f"widgets_gen{g}", f"Widgets (Gen {g})", "hero_widgets")

    data["metadata"]["wave4_updated_utc"] = datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    data["metadata"].setdefault("notes", []).append(
        "Wave 4: pet max levels/skills/refinement + troop tiers + svs rates + expert xp rate; wostools.net cross-reference"
    )

    GAME_DATA.write_text(json.dumps(data, indent=1), encoding="utf-8")
    counts = {k: (len(v) if isinstance(v, list) else "obj") for k, v in data.items() if k in (
        "pet_levels", "troop_tiers", "resource_types")}
    print("written", counts)
    print("gens", gens)
    print("pet max levels", max_levels)


main()
             